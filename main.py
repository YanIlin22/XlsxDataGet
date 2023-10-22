from openpyxl import load_workbook
import tkinter
from tkinter import PhotoImage
from tkinter import ttk
from tkinter.messagebox import showerror
from tkinter.filedialog import asksaveasfile, askopenfile

filename = ''

error = False

noPhoneNumbers = ""

group = ""

def open_file():
    global filename
    i = askopenfile(mode="r")
    if i is None:
        return
    filename = i.name
    #print(filename)
    opened_label.config(text="Файл открыт", foreground="green")
    btngen["state"] = tkinter.NORMAL
    btngen_parents["state"] = tkinter.NORMAL
    combobox['state'] = tkinter.NORMAL
    combobox['values'] = get_groups(filename)
    combobox['values'] = (*combobox['values'], "Выберите группу")
    combobox.config(state="readonly")


def gen():
    generate(filename)

def gen_parents():
    generate_parents(filename)

def generate(file):
    global error
    global noPhoneNumbers
    global group
    try:
        workbook = load_workbook(filename=file)

        sheet = workbook.active
        #vcard = open('card.vcf', 'w', encoding='utf-8')
        files = [('VCARD', '*.vcf')]

        vcard = asksaveasfile(mode='w', filetypes=files, defaultextension=".vcf", title="Сохранить файл")
        for i, row in enumerate(sheet.rows):
            if i == 0:
                continue

            l_name = row[1].value
            name = row[2].value
            m_name = row[3].value
            phone = row[36].value
            current_group = row[20].value

            if l_name is None:
                continue

            if combobox.get() == "" or combobox.get() == "Выберите группу":
                showerror(title="Выберите группу", message="Выберите группу!")
                error = True
                root.destroy()
                break

            if current_group != combobox.get():
                continue

            if phone is None or phone == "-":
                # phone_formatted = "---"
                noPhoneNumbers = noPhoneNumbers + l_name +" "+ name +" "+ m_name+ "\n"
                continue
            else:
                phone_formatted = phone[0] + ' ' + phone[1:4] + ' ' + phone[4:7] + '-' + phone[7:9] + '-' + phone[9:12]
            try:
                with open(vcard, 'w', encoding='utf-8') as file:
                    file.write(
                        f'BEGIN:VCARD\nVERSION:3.0\nFN:{l_name} {name} {m_name}\nitem1.TEL:{phone_formatted}\n'
                        f'item1.X-ABLabel:'f'\nCATEGORIES:Категория\nEND:VCARD\n')
            except Exception:
                vcard.close()
    except Exception:
        showerror(title="Ошибка", message="Произошла ошибка!")
        error = True
        root.destroy()
    try:
        if noPhoneNumbers == "":
            if error is False:
                tkinter.messagebox.showinfo(title="Сформировано!", message="Успешно сформировано!")
                root.destroy()
        else:
            if error is False:
                tkinter.messagebox.showinfo(title="Сформировано!", message="Успешно сформировано!"
                                                                           " Номера телефонов не "
                                                                           "найдены: " + noPhoneNumbers)
            root.destroy()
    except Exception:
        #showerror(title="Ошибка", message="Не сформировано")
        root.destroy()

def generate_parents(file):
    global error
    global noPhoneNumbers
    global group
    try:
        workbook = load_workbook(filename=file)

        sheet = workbook.active
        files = [('VCARD', '*.vcf')]

        vcard = asksaveasfile(mode='w', filetypes=files, defaultextension=".vcf", title="Сохранить файл")

        for i, row in enumerate(sheet.rows):
            if i == 0:
                continue

            fio = row[38].value

            if fio is None:
                continue

            if combobox.get() == "" or combobox.get() == "Выберите группу":
                showerror(title="Выберите группу", message="Выберите группу!")
                error = True
                root.destroy()
                break

            fio = fio.split()
            l_name = fio[0]
            name = fio[1]
            m_name = fio[2]
            phone = row[39].value
            current_group = row[20].value



            if current_group != combobox.get():
                continue

            if phone is None or phone == "-":
                # phone_formatted = "---"
                noPhoneNumbers = noPhoneNumbers + l_name +" "+ name +" "+ m_name+ "\n"
                continue
            else:
                phone_formatted = phone[0] + ' ' + phone[1:4] + ' ' + phone[4:7] + '-' + phone[7:9] + '-' + phone[9:12]
            try:
                with open(vcard, 'w', encoding='utf-8'):
                    vcard.write(
                        f'BEGIN:VCARD\nVERSION:3.0\nFN:{l_name} {name} {m_name}\nitem1.TEL:{phone_formatted}\n'
                        f'item1.X-ABLabel:'f'\nCATEGORIES:Категория\nEND:VCARD\n')
            except Exception:
                vcard.close()
    except Exception:
        showerror(title="Ошибка", message="Произошла ошибка!")
        error = True
        root.destroy()
    try:
        if noPhoneNumbers == "":
            if error is False:
                tkinter.messagebox.showinfo(title="Сформировано!", message="Успешно сформировано!")
                root.destroy()
        else:
            if error is False:
                tkinter.messagebox.showinfo(title="Сформировано!", message="Успешно сформировано!"
                                                                           " Номера телефонов не "
                                                                           "найдены: " + noPhoneNumbers)
            root.destroy()
    except Exception:
        #showerror(title="Ошибка", message="Не сформировано")
        root.destroy()
def get_groups(file):
    global error
    try:
        workbook = load_workbook(filename=file)

        sheet = workbook.active
        groups = set()

        for i, row in enumerate(sheet.rows):
            if i == 0:
                continue

            group = row[20].value
            if group is None:
                continue
            #print(group)
            groups.add(group)

        return sorted(list(groups))
    except Exception:
        showerror(title="Ошибка", message="Произошла ошибка. Проверьте формат файла!")
        error = True


def group_value(i):
    global group
    group = combobox.get()
root = tkinter.Tk()
root.title('XlsxDataGet')
root.geometry("250x235")
root.resizable(False, False)
open_label = tkinter.Label(text="Откройте файл Xlsx")
open_label.grid(sticky="n", padx=2 ,pady=2, column=0)
btnopen = ttk.Button(text="Открыть", command=open_file)
btnopen.grid(sticky="n",ipadx=75, ipady=10,padx=12 , row=2)
opened_label = tkinter.Label(text="Файл не открыт")
opened_label.grid(padx=2 ,pady=3)
combobox = ttk.Combobox(state=tkinter.DISABLED)
combobox.config(state="readonly")
combobox.set("Выберите группу")
combobox.grid(padx=2, ipadx=35)
combobox.bind("<<ComboboxSelected>>", group_value)
#ttk.Entry().pack()
btngen = ttk.Button(text="Ученики", state=tkinter.DISABLED, command=gen)
btngen.grid(pady=7, padx=12 ,ipadx=75, ipady=10)
btngen_parents = ttk.Button(text="Родители", state=tkinter.DISABLED, command=gen_parents)
btngen_parents.grid(sticky="s", padx=12, row=6 ,ipadx=75, ipady=10)
icon = PhotoImage(file='icon.png')
root.iconphoto(False, icon)
root.mainloop()
